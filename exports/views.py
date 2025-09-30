from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
import io

from projects.models import Project
from projects.utils import annotate_year_hours
from xhtml2pdf import pisa

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import date
from projects.models import Project
from timesheets.models import TimesheetEntry
from django.db.models import Sum, Count, Q

# Create your views here.

def _filtered_projects(request):
    qs = Project.objects.all().order_by('-date_received')
    q = request.GET.get('q','').strip()
    status = request.GET.get('status','').strip()
    if q:
        from django.db.models import Q
        qs = qs.filter(Q(number__icontains=q) | Q(name__icontains=q) | Q(client_name__icontains=q))
    if status:
        qs = qs.filter(status=status)
    return annotate_year_hours(qs)

def projects_table_pdf(request):
    projects = _filtered_projects(request)
    html = render_to_string("exports/projects_table_pdf.html", {"projects": projects})
    # render HTML -> PDF (bytes)
    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.StringIO(html), dest=result, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse("Błąd generowania PDF", status=500)
    pdf = result.getvalue()
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp['Content-Disposition'] = 'inline; filename="projekty.pdf"'
    return resp


def projects_table_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projekty"

    # --- Arkusz 1 (istniejący) ---
    headers = ["Nr", "D-nr", "Nazwa", "Zlecający", "Status", "Data przyjęcia", "Termin"]
    ws.append(headers)
    for p in Project.objects.all().order_by("date_received"):
        ws.append([p.number, p.d_number, p.name, p.client_name, p.get_status_display(), p.date_received, p.date_due])

    for col_idx in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # --- Arkusz 2: Godziny per miesiąc (curr year) ---
    curr_year = date.today().year
    ws2 = wb.create_sheet(title=f"Godziny {curr_year}")

    months = ["I","II","III","IV","V","VI","VII","VIII","IX","X","XI","XII"]
    ws2.append(["Projekt"] + months + ["Suma"])

    projects = Project.objects.all().order_by("number")
    for p in projects:
        row = [f"{p.number} {p.name}"]
        monthly = [0] * 12
        qs = (TimesheetEntry.objects
              .filter(project=p, date__year=curr_year, state=TimesheetEntry.State.APPROVED)
              .values_list('date__month')
              .annotate(h=Sum('hours')))
        for m, h in qs:
            monthly[m-1] = float(h)
        row += monthly + [sum(monthly)]
        ws2.append(row)

    for col_idx in range(1, 1 + 12 + 2):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 14

    # --- Response ---
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = f'attachment; filename="projekty_{date.today()}.xlsx"'
    wb.save(resp)
    return resp
